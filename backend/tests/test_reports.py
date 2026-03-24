"""Integration tests for Phase 13 — Reports module."""

from __future__ import annotations

from datetime import date, datetime, timezone
from decimal import Decimal
from io import BytesIO
import re

import pytest
from openpyxl import load_workbook
from werkzeug.security import generate_password_hash

from app.extensions import db as _db
from app.models.annual_quota import AnnualQuota
from app.models.article import Article
from app.models.article_supplier import ArticleSupplier
from app.models.batch import Batch
from app.models.category import Category
from app.models.employee import Employee
from app.models.enums import QuotaEnforcement, TxType, UserRole
from app.models.location import Location
from app.models.personal_issuance import PersonalIssuance
from app.models.receiving import Receiving
from app.models.stock import Stock
from app.models.supplier import Supplier
from app.models.surplus import Surplus
from app.models.transaction import Transaction
from app.models.uom_catalog import UomCatalog
from app.models.user import User
from app.services import report_service


FROZEN_NOW = datetime(2026, 3, 14, 12, 0, tzinfo=timezone.utc)
REPORT_DATE_FROM = "2026-03-01"
REPORT_DATE_TO = "2026-03-31"
REPORT_GENERAL_CATEGORY = "rep13_general"
REPORT_PPE_CATEGORY = "rep13_ppe"
SEEDED_MARCH_INBOUND = Decimal("1260.000")
SEEDED_MARCH_OUTBOUND = Decimal("1020.000")
SEEDED_REORDER_INCREMENT = {"RED": 1, "YELLOW": 1, "NORMAL": 2}
_OUTBOUND_TYPES = {
    TxType.OUTBOUND,
    TxType.STOCK_CONSUMED,
    TxType.SURPLUS_CONSUMED,
    TxType.PERSONAL_ISSUE,
}
_FILENAME_RE = re.compile(r'filename="?([^";]+)"?')


class _FrozenReportDateTime(datetime):
    """Freeze report-service time-sensitive windows for deterministic tests."""

    @classmethod
    def now(cls, tz=None):
        base = cls(
            FROZEN_NOW.year,
            FROZEN_NOW.month,
            FROZEN_NOW.day,
            FROZEN_NOW.hour,
            FROZEN_NOW.minute,
            FROZEN_NOW.second,
            tzinfo=timezone.utc,
        )
        if tz is None:
            return base
        return base.astimezone(tz)


def _decimal(value) -> Decimal:
    return Decimal(str(value))


def _movement_totals_for_month(month_start: date) -> dict[str, float]:
    next_month = (
        date(month_start.year + 1, 1, 1)
        if month_start.month == 12
        else date(month_start.year, month_start.month + 1, 1)
    )
    started_at = datetime(month_start.year, month_start.month, month_start.day, tzinfo=timezone.utc)
    ended_at = datetime(next_month.year, next_month.month, next_month.day, tzinfo=timezone.utc)

    inbound = Decimal("0")
    outbound = Decimal("0")
    rows = (
        Transaction.query
        .with_entities(Transaction.tx_type, Transaction.quantity)
        .filter(
            Transaction.occurred_at >= started_at,
            Transaction.occurred_at < ended_at,
        )
        .all()
    )
    for tx_type, quantity in rows:
        amount = _decimal(quantity).copy_abs()
        if tx_type == TxType.STOCK_RECEIPT:
            inbound += amount
        elif tx_type in _OUTBOUND_TYPES:
            outbound += amount

    return {"inbound": float(inbound), "outbound": float(outbound)}


def _extract_filename(content_disposition: str) -> str:
    match = _FILENAME_RE.search(content_disposition)
    assert match, content_disposition
    return match.group(1)


def _item_by_article(items: list[dict], article_no: str) -> dict:
    return next(item for item in items if item["article_no"] == article_no)


@pytest.fixture(autouse=True)
def freeze_report_time(monkeypatch):
    monkeypatch.setattr(report_service, "datetime", _FrozenReportDateTime)


# ---------------------------------------------------------------------------
# Module-scoped fixtures
# ---------------------------------------------------------------------------


@pytest.fixture(scope="module")
def reports_data(app):
    """Seed Reports-specific fixtures once per module."""
    with app.app_context():
        baseline_reorder = {
            item["reorder_status"]: item["count"]
            for item in report_service.get_reorder_summary_statistics()["items"]
        }
        baseline_march = _movement_totals_for_month(date(2026, 3, 1))

        location = _db.session.get(Location, 1)
        if location is None:
            location = Location(id=1, name="Reports Warehouse", timezone="UTC", is_active=True)
            _db.session.add(location)
            _db.session.flush()

        kg = UomCatalog.query.filter_by(code="rep13_kg").first()
        if kg is None:
            kg = UomCatalog(code="rep13_kg", label_hr="kilogram", decimal_display=True)
            _db.session.add(kg)
            _db.session.flush()

        kom = UomCatalog.query.filter_by(code="rep13_kom").first()
        if kom is None:
            kom = UomCatalog(code="rep13_kom", label_hr="komad", decimal_display=False)
            _db.session.add(kom)
            _db.session.flush()

        general_category = Category.query.filter_by(key=REPORT_GENERAL_CATEGORY).first()
        if general_category is None:
            general_category = Category(
                key=REPORT_GENERAL_CATEGORY,
                label_hr="Reports General",
                is_active=True,
            )
            _db.session.add(general_category)
            _db.session.flush()

        ppe_category = Category.query.filter_by(key=REPORT_PPE_CATEGORY).first()
        if ppe_category is None:
            ppe_category = Category(
                key=REPORT_PPE_CATEGORY,
                label_hr="Reports PPE",
                is_personal_issue=True,
                is_active=True,
            )
            _db.session.add(ppe_category)
            _db.session.flush()

        supplier_primary = Supplier.query.filter_by(internal_code="REP13-SUP-001").first()
        if supplier_primary is None:
            supplier_primary = Supplier(
                internal_code="REP13-SUP-001",
                name="Reports Preferred Supplier",
                is_active=True,
            )
            _db.session.add(supplier_primary)
            _db.session.flush()

        supplier_secondary = Supplier.query.filter_by(internal_code="REP13-SUP-002").first()
        if supplier_secondary is None:
            supplier_secondary = Supplier(
                internal_code="REP13-SUP-002",
                name="Reports Secondary Supplier",
                is_active=True,
            )
            _db.session.add(supplier_secondary)
            _db.session.flush()

        admin = User.query.filter_by(username="rep13_admin").first()
        if admin is None:
            admin = User(
                username="rep13_admin",
                password_hash=generate_password_hash("adminpass", method="pbkdf2:sha256"),
                role=UserRole.ADMIN,
                is_active=True,
            )
            _db.session.add(admin)

        manager = User.query.filter_by(username="rep13_manager").first()
        if manager is None:
            manager = User(
                username="rep13_manager",
                password_hash=generate_password_hash("managerpass", method="pbkdf2:sha256"),
                role=UserRole.MANAGER,
                is_active=True,
            )
            _db.session.add(manager)

        _db.session.flush()

        article_yellow = Article.query.filter_by(article_no="REP13-001").first()
        if article_yellow is None:
            article_yellow = Article(
                article_no="REP13-001",
                description="Yellow reorder article",
                category_id=general_category.id,
                base_uom=kg.id,
                has_batch=False,
                reorder_threshold=Decimal("10.000"),
                is_active=True,
            )
            _db.session.add(article_yellow)
            _db.session.flush()

        article_red = Article.query.filter_by(article_no="REP13-002").first()
        if article_red is None:
            article_red = Article(
                article_no="REP13-002",
                description="Red reorder article",
                category_id=general_category.id,
                base_uom=kg.id,
                has_batch=True,
                reorder_threshold=Decimal("5.000"),
                is_active=True,
            )
            _db.session.add(article_red)
            _db.session.flush()

        article_normal = Article.query.filter_by(article_no="REP13-003").first()
        if article_normal is None:
            article_normal = Article(
                article_no="REP13-003",
                description="Normal reorder article",
                category_id=general_category.id,
                base_uom=kg.id,
                has_batch=True,
                reorder_threshold=Decimal("5.000"),
                is_active=True,
            )
            _db.session.add(article_normal)
            _db.session.flush()

        article_ppe = Article.query.filter_by(article_no="REP13-004").first()
        if article_ppe is None:
            article_ppe = Article(
                article_no="REP13-004",
                description="Personal issue article",
                category_id=ppe_category.id,
                base_uom=kom.id,
                has_batch=False,
                reorder_threshold=Decimal("1.000"),
                is_active=True,
            )
            _db.session.add(article_ppe)
            _db.session.flush()

        article_stats = Article.query.filter_by(article_no="REP13-099").first()
        if article_stats is None:
            article_stats = Article(
                article_no="REP13-099",
                description="Statistics leader article",
                category_id=general_category.id,
                base_uom=kg.id,
                has_batch=False,
                reorder_threshold=Decimal("1.000"),
                is_active=False,
            )
            _db.session.add(article_stats)
            _db.session.flush()

        for supplier_id, is_preferred in (
            (supplier_secondary.id, False),
            (supplier_primary.id, True),
        ):
            link = ArticleSupplier.query.filter_by(
                article_id=article_yellow.id,
                supplier_id=supplier_id,
            ).first()
            if link is None:
                _db.session.add(
                    ArticleSupplier(
                        article_id=article_yellow.id,
                        supplier_id=supplier_id,
                        is_preferred=is_preferred,
                    )
                )

        red_batch = Batch.query.filter_by(batch_code="REP13-B002", article_id=article_red.id).first()
        if red_batch is None:
            red_batch = Batch(
                article_id=article_red.id,
                batch_code="REP13-B002",
                expiry_date=date(2026, 9, 30),
            )
            _db.session.add(red_batch)
            _db.session.flush()

        normal_batch = Batch.query.filter_by(
            batch_code="REP13-B003",
            article_id=article_normal.id,
        ).first()
        if normal_batch is None:
            normal_batch = Batch(
                article_id=article_normal.id,
                batch_code="REP13-B003",
                expiry_date=date(2026, 12, 31),
            )
            _db.session.add(normal_batch)
            _db.session.flush()

        stock_yellow = Stock.query.filter_by(
            location_id=location.id,
            article_id=article_yellow.id,
            batch_id=None,
        ).first()
        if stock_yellow is None:
            _db.session.add(
                Stock(
                    location_id=location.id,
                    article_id=article_yellow.id,
                    batch_id=None,
                    quantity=Decimal("8.000"),
                    uom=kg.code,
                    average_price=Decimal("5.0000"),
                )
            )

        stock_red = Stock.query.filter_by(
            location_id=location.id,
            article_id=article_red.id,
            batch_id=red_batch.id,
        ).first()
        if stock_red is None:
            _db.session.add(
                Stock(
                    location_id=location.id,
                    article_id=article_red.id,
                    batch_id=red_batch.id,
                    quantity=Decimal("4.000"),
                    uom=kg.code,
                    average_price=Decimal("7.0000"),
                )
            )

        stock_normal = Stock.query.filter_by(
            location_id=location.id,
            article_id=article_normal.id,
            batch_id=normal_batch.id,
        ).first()
        if stock_normal is None:
            _db.session.add(
                Stock(
                    location_id=location.id,
                    article_id=article_normal.id,
                    batch_id=normal_batch.id,
                    quantity=Decimal("25.000"),
                    uom=kg.code,
                    average_price=Decimal("4.0000"),
                )
            )

        stock_ppe = Stock.query.filter_by(
            location_id=location.id,
            article_id=article_ppe.id,
            batch_id=None,
        ).first()
        if stock_ppe is None:
            _db.session.add(
                Stock(
                    location_id=location.id,
                    article_id=article_ppe.id,
                    batch_id=None,
                    quantity=Decimal("20.000"),
                    uom=kom.code,
                    average_price=Decimal("3.0000"),
                )
            )

        surplus_yellow = Surplus.query.filter_by(
            location_id=location.id,
            article_id=article_yellow.id,
            batch_id=None,
        ).first()
        if surplus_yellow is None:
            _db.session.add(
                Surplus(
                    location_id=location.id,
                    article_id=article_yellow.id,
                    batch_id=None,
                    quantity=Decimal("3.000"),
                    uom=kg.code,
                    created_at=datetime(2026, 3, 12, 10, 0, tzinfo=timezone.utc),
                )
            )

        surplus_normal = Surplus.query.filter_by(
            location_id=location.id,
            article_id=article_normal.id,
            batch_id=normal_batch.id,
        ).first()
        if surplus_normal is None:
            _db.session.add(
                Surplus(
                    location_id=location.id,
                    article_id=article_normal.id,
                    batch_id=normal_batch.id,
                    quantity=Decimal("6.000"),
                    uom=kg.code,
                    created_at=datetime(2026, 3, 13, 9, 30, tzinfo=timezone.utc),
                )
            )

        employee = Employee.query.filter_by(employee_id="REP13-EMP-001").first()
        if employee is None:
            employee = Employee(
                employee_id="REP13-EMP-001",
                first_name="Ivo",
                last_name="Ivic",
                job_title="Reports Painter",
                is_active=True,
            )
            _db.session.add(employee)
            _db.session.flush()

        quota = (
            AnnualQuota.query
            .filter_by(
                job_title="Reports Painter",
                category_id=ppe_category.id,
                article_id=None,
                employee_id=None,
            )
            .first()
        )
        if quota is None:
            _db.session.add(
                AnnualQuota(
                    job_title="Reports Painter",
                    category_id=ppe_category.id,
                    article_id=None,
                    employee_id=None,
                    quantity=Decimal("10.000"),
                    uom=kom.code,
                    reset_month=1,
                    enforcement=QuotaEnforcement.WARN,
                )
            )

        for issued_at, quantity in (
            (datetime(2026, 3, 10, 8, 0, tzinfo=timezone.utc), Decimal("4.000")),
            (datetime(2025, 12, 30, 8, 0, tzinfo=timezone.utc), Decimal("2.000")),
        ):
            issuance = PersonalIssuance.query.filter_by(
                employee_id=employee.id,
                article_id=article_ppe.id,
                issued_at=issued_at,
            ).first()
            if issuance is None:
                _db.session.add(
                    PersonalIssuance(
                        employee_id=employee.id,
                        article_id=article_ppe.id,
                        batch_id=None,
                        quantity=quantity,
                        uom=kom.code,
                        issued_by=admin.id,
                        issued_at=issued_at,
                    )
                )

        transactions = [
            {
                "article_id": article_yellow.id,
                "batch_id": None,
                "tx_type": TxType.STOCK_RECEIPT,
                "occurred_at": datetime(2026, 3, 1, 9, 0, tzinfo=timezone.utc),
                "quantity": Decimal("50.000"),
                "uom": kg.code,
                "reference_type": "receiving",
                "reference_id": 5101,
                "delivery_note_number": "REP13-DN-001",
            },
            {
                "article_id": article_yellow.id,
                "batch_id": None,
                "tx_type": TxType.STOCK_CONSUMED,
                "occurred_at": datetime(2026, 3, 5, 12, 0, tzinfo=timezone.utc),
                "quantity": Decimal("-12.000"),
                "uom": kg.code,
                "reference_type": "draft",
                "reference_id": 6101,
            },
            {
                "article_id": article_yellow.id,
                "batch_id": None,
                "tx_type": TxType.SURPLUS_CONSUMED,
                "occurred_at": datetime(2026, 3, 6, 12, 0, tzinfo=timezone.utc),
                "quantity": Decimal("-3.000"),
                "uom": kg.code,
                "reference_type": "draft",
                "reference_id": 6102,
            },
            {
                "article_id": article_yellow.id,
                "batch_id": None,
                "tx_type": TxType.OUTBOUND,
                "occurred_at": datetime(2026, 3, 7, 12, 0, tzinfo=timezone.utc),
                "quantity": Decimal("-1.000"),
                "uom": kg.code,
                "reference_type": "draft",
                "reference_id": 6103,
            },
            {
                "article_id": article_red.id,
                "batch_id": red_batch.id,
                "tx_type": TxType.STOCK_RECEIPT,
                "occurred_at": datetime(2026, 3, 2, 9, 0, tzinfo=timezone.utc),
                "quantity": Decimal("10.000"),
                "uom": kg.code,
                "reference_type": "receiving",
                "reference_id": 5102,
                "delivery_note_number": "REP13-DN-002",
            },
            {
                "article_id": article_ppe.id,
                "batch_id": None,
                "tx_type": TxType.PERSONAL_ISSUE,
                "occurred_at": datetime(2026, 3, 10, 8, 0, tzinfo=timezone.utc),
                "quantity": Decimal("-4.000"),
                "uom": kom.code,
                "reference_type": "issuance",
                "reference_id": 7101,
            },
            {
                "article_id": article_normal.id,
                "batch_id": normal_batch.id,
                "tx_type": TxType.INVENTORY_ADJUSTMENT,
                "occurred_at": datetime(2026, 3, 11, 9, 0, tzinfo=timezone.utc),
                "quantity": Decimal("2.000"),
                "uom": kg.code,
                "reference_type": "inventory_count",
                "reference_id": 8101,
            },
            {
                "article_id": article_stats.id,
                "batch_id": None,
                "tx_type": TxType.STOCK_RECEIPT,
                "occurred_at": datetime(2026, 3, 3, 10, 0, tzinfo=timezone.utc),
                "quantity": Decimal("1200.000"),
                "uom": kg.code,
                "reference_type": "receiving",
                "reference_id": 5103,
                "delivery_note_number": "REP13-DN-099",
            },
            {
                "article_id": article_stats.id,
                "batch_id": None,
                "tx_type": TxType.OUTBOUND,
                "occurred_at": datetime(2026, 3, 4, 11, 0, tzinfo=timezone.utc),
                "quantity": Decimal("-1000.000"),
                "uom": kg.code,
                "reference_type": "draft",
                "reference_id": 6104,
            },
        ]

        for payload in transactions:
            existing = (
                Transaction.query
                .filter_by(
                    article_id=payload["article_id"],
                    tx_type=payload["tx_type"],
                    occurred_at=payload["occurred_at"],
                    reference_type=payload["reference_type"],
                    reference_id=payload["reference_id"],
                )
                .first()
            )
            if existing is None:
                tx_payload = dict(payload)
                _db.session.add(
                    Transaction(
                        location_id=location.id,
                        user_id=admin.id,
                        delivery_note_number=tx_payload.pop("delivery_note_number", None),
                        **tx_payload,
                    )
                )

        _db.session.commit()

        return {
            "admin_username": "rep13_admin",
            "admin_password": "adminpass",
            "manager_username": "rep13_manager",
            "manager_password": "managerpass",
            "article_yellow_id": article_yellow.id,
            "article_red_id": article_red.id,
            "article_normal_id": article_normal.id,
            "article_ppe_id": article_ppe.id,
            "article_stats_id": article_stats.id,
            "baseline_reorder": baseline_reorder,
            "baseline_march": baseline_march,
        }


# ---------------------------------------------------------------------------
# Auth helpers
# ---------------------------------------------------------------------------


_token_cache: dict[str, str] = {}


def _get_token(client, username: str, password: str, ip: str) -> str:
    cache_key = f"{username}:{ip}"
    if cache_key not in _token_cache:
        response = client.post(
            "/api/v1/auth/login",
            json={"username": username, "password": password},
            environ_base={"REMOTE_ADDR": ip},
        )
        assert response.status_code == 200, response.get_json()
        _token_cache[cache_key] = response.get_json()["access_token"]
    return _token_cache[cache_key]


def _admin_headers(client, reports_data):
    token = _get_token(
        client,
        reports_data["admin_username"],
        reports_data["admin_password"],
        "127.0.33.1",
    )
    return {"Authorization": f"Bearer {token}"}


def _manager_headers(client, reports_data):
    token = _get_token(
        client,
        reports_data["manager_username"],
        reports_data["manager_password"],
        "127.0.33.2",
    )
    return {"Authorization": f"Bearer {token}"}


# ---------------------------------------------------------------------------
# Stock overview tests
# ---------------------------------------------------------------------------


def test_stock_overview_returns_seeded_metrics_for_known_period(client, reports_data):
    response = client.get(
        "/api/v1/reports/stock-overview"
        f"?date_from={REPORT_DATE_FROM}&date_to={REPORT_DATE_TO}&category={REPORT_GENERAL_CATEGORY}",
        headers=_admin_headers(client, reports_data),
    )

    assert response.status_code == 200, response.get_json()
    payload = response.get_json()
    assert payload["period"] == {
        "date_from": REPORT_DATE_FROM,
        "date_to": REPORT_DATE_TO,
        "months": 0.99,
    }
    assert payload["total"] == 3

    yellow_item = _item_by_article(payload["items"], "REP13-001")
    assert yellow_item["supplier_name"] == "Reports Preferred Supplier"
    assert yellow_item["stock"] == 8.0
    assert yellow_item["surplus"] == 3.0
    assert yellow_item["total_available"] == 11.0
    assert yellow_item["inbound"] == 50.0
    assert yellow_item["outbound"] == 16.0
    assert yellow_item["avg_monthly_consumption"] == 16.16
    assert yellow_item["coverage_months"] == 0.7
    assert yellow_item["reorder_threshold"] == 10.0
    assert yellow_item["reorder_status"] == "YELLOW"

    red_item = _item_by_article(payload["items"], "REP13-002")
    assert red_item["stock"] == 4.0
    assert red_item["surplus"] == 0.0
    assert red_item["total_available"] == 4.0
    assert red_item["inbound"] == 10.0
    assert red_item["outbound"] == 0.0
    assert red_item["avg_monthly_consumption"] == 0.0
    assert red_item["coverage_months"] is None
    assert red_item["reorder_status"] == "RED"

    normal_item = _item_by_article(payload["items"], "REP13-003")
    assert normal_item["stock"] == 25.0
    assert normal_item["surplus"] == 6.0
    assert normal_item["total_available"] == 31.0
    assert normal_item["avg_monthly_consumption"] == 0.0
    assert normal_item["coverage_months"] is None
    assert normal_item["reorder_status"] == "NORMAL"


def test_stock_overview_reorder_status_matches_locked_warehouse_semantics(client, reports_data):
    response = client.get(
        "/api/v1/reports/stock-overview"
        f"?date_from={REPORT_DATE_FROM}&date_to={REPORT_DATE_TO}"
        f"&category={REPORT_GENERAL_CATEGORY}&reorder_only=true",
        headers=_admin_headers(client, reports_data),
    )

    assert response.status_code == 200, response.get_json()
    payload = response.get_json()
    assert payload["total"] == 2
    assert [item["article_no"] for item in payload["items"]] == ["REP13-001", "REP13-002"]
    assert [item["reorder_status"] for item in payload["items"]] == ["YELLOW", "RED"]


def test_stock_overview_response_includes_value_fields_and_summary(client, reports_data):
    """Value fields and summary key exist in a standard response (no price data seeded here)."""
    response = client.get(
        "/api/v1/reports/stock-overview"
        f"?date_from={REPORT_DATE_FROM}&date_to={REPORT_DATE_TO}&category={REPORT_GENERAL_CATEGORY}",
        headers=_admin_headers(client, reports_data),
    )

    assert response.status_code == 200, response.get_json()
    payload = response.get_json()
    assert "summary" in payload
    assert "warehouse_total_value" in payload["summary"]
    for item in payload["items"]:
        assert "unit_value" in item
        assert "total_value" in item


# ---------------------------------------------------------------------------
# Stock overview value contract fixtures and tests
# ---------------------------------------------------------------------------


@pytest.fixture(scope="module")
def reports_value_data(app, reports_data):
    """Seed Receiving rows and supplier last_price for value contract tests.

    article_yellow (REP13-001):
      - preferred supplier last_price = 15.00  (should lose to receiving price)
      - older Receiving with unit_price = 12.50  (2026-01-01, should lose to newer)
      - newer Receiving with unit_price = 20.00  (2026-03-01, SHOULD WIN)
      - newest Receiving with unit_price = NULL   (2026-03-20, must NOT erase 20.00)
      -> expected unit_value = 20.00, total_value = 8.0 * 20.00 = 160.00

    article_normal (REP13-003):
      - no Receiving rows with price
      - preferred supplier last_price = 8.75  (should be used as fallback)
      -> expected unit_value = 8.75, total_value = 25.0 * 8.75 = 218.75

    article_red (REP13-002):
      - no Receiving rows, no preferred supplier link
      -> expected unit_value = null, total_value = null

    warehouse_total_value = 160.00 + 218.75 = 378.75
    """
    with app.app_context():
        article_yellow = Article.query.filter_by(article_no="REP13-001").first()
        article_normal = Article.query.filter_by(article_no="REP13-003").first()
        admin = User.query.filter_by(username="rep13_admin").first()
        location = _db.session.get(Location, 1)
        kg = UomCatalog.query.filter_by(code="rep13_kg").first()
        supplier_primary = Supplier.query.filter_by(internal_code="REP13-SUP-001").first()

        # Set preferred supplier last_price for article_yellow
        link_yellow = ArticleSupplier.query.filter_by(
            article_id=article_yellow.id,
            supplier_id=supplier_primary.id,
            is_preferred=True,
        ).first()
        link_yellow.last_price = Decimal("15.00")

        # Add preferred supplier link for article_normal with last_price
        link_normal = ArticleSupplier.query.filter_by(
            article_id=article_normal.id,
            supplier_id=supplier_primary.id,
        ).first()
        if link_normal is None:
            _db.session.add(
                ArticleSupplier(
                    article_id=article_normal.id,
                    supplier_id=supplier_primary.id,
                    is_preferred=True,
                    last_price=Decimal("8.75"),
                )
            )
        else:
            link_normal.is_preferred = True
            link_normal.last_price = Decimal("8.75")

        # Receiving: oldest non-null price for article_yellow (should lose to row 2)
        if not Receiving.query.filter_by(
            article_id=article_yellow.id,
            delivery_note_number="REP13-VAL-DN-001",
        ).first():
            _db.session.add(
                Receiving(
                    article_id=article_yellow.id,
                    location_id=location.id,
                    quantity=Decimal("10.000"),
                    uom=kg.code,
                    unit_price=Decimal("12.50"),
                    delivery_note_number="REP13-VAL-DN-001",
                    received_by=admin.id,
                    received_at=datetime(2026, 1, 1, 10, 0, tzinfo=timezone.utc),
                )
            )

        # Receiving: newer non-null price for article_yellow (SHOULD WIN)
        if not Receiving.query.filter_by(
            article_id=article_yellow.id,
            delivery_note_number="REP13-VAL-DN-002",
        ).first():
            _db.session.add(
                Receiving(
                    article_id=article_yellow.id,
                    location_id=location.id,
                    quantity=Decimal("20.000"),
                    uom=kg.code,
                    unit_price=Decimal("20.00"),
                    delivery_note_number="REP13-VAL-DN-002",
                    received_by=admin.id,
                    received_at=datetime(2026, 3, 1, 10, 0, tzinfo=timezone.utc),
                )
            )

        # Receiving: newest but null price for article_yellow (must NOT erase row 2)
        if not Receiving.query.filter_by(
            article_id=article_yellow.id,
            delivery_note_number="REP13-VAL-DN-003",
        ).first():
            _db.session.add(
                Receiving(
                    article_id=article_yellow.id,
                    location_id=location.id,
                    quantity=Decimal("5.000"),
                    uom=kg.code,
                    unit_price=None,
                    delivery_note_number="REP13-VAL-DN-003",
                    received_by=admin.id,
                    received_at=datetime(2026, 3, 20, 10, 0, tzinfo=timezone.utc),
                )
            )

        _db.session.commit()


def test_stock_overview_value_receiving_price_wins_over_supplier_last_price(
    client, reports_data, reports_value_data
):
    """Most recent non-null receiving price beats preferred supplier last_price."""
    response = client.get(
        "/api/v1/reports/stock-overview"
        f"?date_from={REPORT_DATE_FROM}&date_to={REPORT_DATE_TO}&category={REPORT_GENERAL_CATEGORY}",
        headers=_admin_headers(client, reports_data),
    )

    assert response.status_code == 200, response.get_json()
    payload = response.get_json()
    yellow_item = _item_by_article(payload["items"], "REP13-001")
    assert yellow_item["unit_value"] == 20.0


def test_stock_overview_value_null_receiving_does_not_erase_older_known_price(
    client, reports_data, reports_value_data
):
    """A newer Receiving row with unit_price=NULL must not overwrite an older non-null price."""
    response = client.get(
        "/api/v1/reports/stock-overview"
        f"?date_from={REPORT_DATE_FROM}&date_to={REPORT_DATE_TO}&category={REPORT_GENERAL_CATEGORY}",
        headers=_admin_headers(client, reports_data),
    )

    assert response.status_code == 200, response.get_json()
    payload = response.get_json()
    yellow_item = _item_by_article(payload["items"], "REP13-001")
    # Must be 20.00 (non-null 2026-03-01 row), not null (2026-03-20 null row)
    assert yellow_item["unit_value"] == 20.0


def test_stock_overview_value_preferred_supplier_fallback_when_no_receiving_price(
    client, reports_data, reports_value_data
):
    """When no receiving price exists, preferred supplier last_price is used."""
    response = client.get(
        "/api/v1/reports/stock-overview"
        f"?date_from={REPORT_DATE_FROM}&date_to={REPORT_DATE_TO}&category={REPORT_GENERAL_CATEGORY}",
        headers=_admin_headers(client, reports_data),
    )

    assert response.status_code == 200, response.get_json()
    payload = response.get_json()
    normal_item = _item_by_article(payload["items"], "REP13-003")
    assert normal_item["unit_value"] == 8.75
    assert normal_item["total_value"] == round(25.0 * 8.75, 2)


def test_stock_overview_value_null_when_no_price_data(
    client, reports_data, reports_value_data
):
    """Article with no receiving price and no preferred-supplier last_price returns nulls."""
    response = client.get(
        "/api/v1/reports/stock-overview"
        f"?date_from={REPORT_DATE_FROM}&date_to={REPORT_DATE_TO}&category={REPORT_GENERAL_CATEGORY}",
        headers=_admin_headers(client, reports_data),
    )

    assert response.status_code == 200, response.get_json()
    payload = response.get_json()
    red_item = _item_by_article(payload["items"], "REP13-002")
    assert red_item["unit_value"] is None
    assert red_item["total_value"] is None


def test_stock_overview_value_total_value_uses_stock_not_surplus(
    client, reports_data, reports_value_data
):
    """total_value = stock × unit_value; surplus is excluded from the calculation."""
    response = client.get(
        "/api/v1/reports/stock-overview"
        f"?date_from={REPORT_DATE_FROM}&date_to={REPORT_DATE_TO}&category={REPORT_GENERAL_CATEGORY}",
        headers=_admin_headers(client, reports_data),
    )

    assert response.status_code == 200, response.get_json()
    payload = response.get_json()
    yellow_item = _item_by_article(payload["items"], "REP13-001")
    # stock=8.0 (not 11.0 which includes surplus 3.0)
    assert yellow_item["stock"] == 8.0
    assert yellow_item["total_value"] == round(8.0 * 20.0, 2)


def test_stock_overview_summary_warehouse_total_excludes_null_price_articles(
    client, reports_data, reports_value_data
):
    """warehouse_total_value sums non-null total_value items; null-price articles are excluded."""
    response = client.get(
        "/api/v1/reports/stock-overview"
        f"?date_from={REPORT_DATE_FROM}&date_to={REPORT_DATE_TO}&category={REPORT_GENERAL_CATEGORY}",
        headers=_admin_headers(client, reports_data),
    )

    assert response.status_code == 200, response.get_json()
    payload = response.get_json()
    # yellow: 8.0 * 20.00 = 160.00
    # normal: 25.0 * 8.75 = 218.75
    # red: null (excluded)
    expected_total = round(8.0 * 20.0 + 25.0 * 8.75, 2)
    assert payload["summary"]["warehouse_total_value"] == expected_total


# ---------------------------------------------------------------------------
# Surplus report tests
# ---------------------------------------------------------------------------


def test_surplus_report_returns_all_seeded_current_surplus_rows(client, reports_data):
    response = client.get(
        "/api/v1/reports/surplus",
        headers=_admin_headers(client, reports_data),
    )

    assert response.status_code == 200, response.get_json()
    payload = response.get_json()
    surplus_by_article = {
        item["article_no"]: item for item in payload["items"] if item["article_no"] in {"REP13-001", "REP13-003"}
    }
    assert set(surplus_by_article) == {"REP13-001", "REP13-003"}

    yellow_row = surplus_by_article["REP13-001"]
    assert yellow_row["surplus_qty"] == 3.0
    assert yellow_row["batch_code"] is None
    assert yellow_row["expiry_date"] is None
    assert yellow_row["discovered"] == "2026-03-12"

    normal_row = surplus_by_article["REP13-003"]
    assert normal_row["surplus_qty"] == 6.0
    assert normal_row["batch_code"] == "REP13-B003"
    assert normal_row["expiry_date"] == "2026-12-31"
    assert normal_row["discovered"] == "2026-03-13"


# ---------------------------------------------------------------------------
# Transaction log tests
# ---------------------------------------------------------------------------


def test_transaction_log_filtered_by_article_returns_only_that_articles_transactions(
    client,
    reports_data,
):
    response = client.get(
        "/api/v1/reports/transactions"
        f"?article_id={reports_data['article_yellow_id']}&page=1&per_page=10",
        headers=_admin_headers(client, reports_data),
    )

    assert response.status_code == 200, response.get_json()
    payload = response.get_json()
    assert payload["total"] == 4
    assert payload["page"] == 1
    assert payload["per_page"] == 10
    assert {item["article_id"] for item in payload["items"]} == {reports_data["article_yellow_id"]}
    assert [item["type"] for item in payload["items"]] == [
        "OUTBOUND",
        "SURPLUS_CONSUMED",
        "STOCK_CONSUMED",
        "STOCK_RECEIPT",
    ]


def test_transaction_log_filtered_by_date_range_returns_only_matching_transactions(
    client,
    reports_data,
):
    response = client.get(
        "/api/v1/reports/transactions"
        f"?article_id={reports_data['article_yellow_id']}"
        "&date_from=2026-03-05&date_to=2026-03-06&page=1&per_page=10",
        headers=_admin_headers(client, reports_data),
    )

    assert response.status_code == 200, response.get_json()
    payload = response.get_json()
    assert payload["total"] == 2
    assert [item["type"] for item in payload["items"]] == [
        "SURPLUS_CONSUMED",
        "STOCK_CONSUMED",
    ]
    assert {item["reference"] for item in payload["items"]} == {"draft:6101", "draft:6102"}


def test_transaction_log_pagination_returns_expected_page_slice(client, reports_data):
    response = client.get(
        "/api/v1/reports/transactions"
        f"?article_id={reports_data['article_yellow_id']}&page=2&per_page=2",
        headers=_admin_headers(client, reports_data),
    )

    assert response.status_code == 200, response.get_json()
    payload = response.get_json()
    assert payload["total"] == 4
    assert payload["page"] == 2
    assert payload["per_page"] == 2
    assert [item["type"] for item in payload["items"]] == ["STOCK_CONSUMED", "STOCK_RECEIPT"]


# ---------------------------------------------------------------------------
# Export tests
# ---------------------------------------------------------------------------


def test_stock_overview_export_xlsx_has_expected_sheet_headers_rows_and_widths(
    client,
    reports_data,
):
    response = client.get(
        "/api/v1/reports/stock-overview/export"
        f"?format=xlsx&date_from={REPORT_DATE_FROM}&date_to={REPORT_DATE_TO}"
        f"&category={REPORT_GENERAL_CATEGORY}",
        headers=_admin_headers(client, reports_data),
    )

    assert response.status_code == 200
    assert response.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert _extract_filename(response.headers["Content-Disposition"]) == "wms_stock_overview_2026-03-14.xlsx"

    workbook = load_workbook(BytesIO(response.data))
    assert workbook.sheetnames == ["Stock Overview"]
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    assert rows[0] == (
        "Article No.",
        "Description",
        "Supplier",
        "Stock",
        "Surplus",
        "Total available",
        "Inbound",
        "Outbound",
        "Avg monthly consumption",
        "Coverage (months)",
        "Reorder threshold",
        "Status",
    )

    rows_by_article = {row[0]: row for row in rows[1:] if row[0]}
    assert rows_by_article["REP13-001"][3:] == (
        "8.0 rep13_kg",
        "3.0 rep13_kg",
        "11.0 rep13_kg",
        "50.0 rep13_kg",
        "16.0 rep13_kg",
        "16.16 rep13_kg",
        0.7,
        "10.0 rep13_kg",
        "YELLOW",
    )
    assert rows_by_article["REP13-002"][9] == "∞"
    assert rows_by_article["REP13-003"][11] == "NORMAL"
    assert sheet.column_dimensions["B"].width >= len("Yellow reorder article") + 1


def test_surplus_export_xlsx_has_expected_sheet_headers_rows_and_widths(client, reports_data):
    response = client.get(
        "/api/v1/reports/surplus/export?format=xlsx",
        headers=_admin_headers(client, reports_data),
    )

    assert response.status_code == 200
    assert response.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert _extract_filename(response.headers["Content-Disposition"]) == "wms_surplus_2026-03-14.xlsx"

    workbook = load_workbook(BytesIO(response.data))
    assert workbook.sheetnames == ["Surplus"]
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    assert rows[0] == (
        "Article No.",
        "Description",
        "Batch",
        "Expiry date",
        "Surplus qty",
        "Discovered",
    )

    rows_by_article = {row[0]: row for row in rows[1:] if row[0]}
    assert rows_by_article["REP13-001"] == (
        "REP13-001",
        "Yellow reorder article",
        "-",
        "-",
        "3.0 rep13_kg",
        "2026-03-12",
    )
    assert rows_by_article["REP13-003"] == (
        "REP13-003",
        "Normal reorder article",
        "REP13-B003",
        "2026-12-31",
        "6.0 rep13_kg",
        "2026-03-13",
    )
    assert sheet.column_dimensions["E"].width >= len("6.0 rep13_kg") + 1


def test_transaction_export_xlsx_has_expected_sheet_headers_rows_and_widths(
    client,
    reports_data,
):
    response = client.get(
        "/api/v1/reports/transactions/export?format=xlsx"
        f"&article_id={reports_data['article_yellow_id']}"
        "&date_from=2026-03-01&date_to=2026-03-31",
        headers=_admin_headers(client, reports_data),
    )

    assert response.status_code == 200
    assert response.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert _extract_filename(response.headers["Content-Disposition"]) == "wms_transactions_2026-03-14.xlsx"

    workbook = load_workbook(BytesIO(response.data))
    assert workbook.sheetnames == ["Transactions"]
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    assert rows[0] == (
        "Occurred at",
        "Article No.",
        "Description",
        "Type",
        "Quantity",
        "Batch",
        "Reference",
        "User",
    )
    assert [row[3] for row in rows[1:]] == [
        "OUTBOUND",
        "SURPLUS_CONSUMED",
        "STOCK_CONSUMED",
        "STOCK_RECEIPT",
    ]
    assert rows[1][4] == "-1.0 rep13_kg"
    assert rows[-1][6] == "REP13-DN-001"
    assert sheet.column_dimensions["A"].width >= len("Occurred at") + 8


@pytest.mark.parametrize(
    ("path", "expected_filename"),
    [
        (
            f"/api/v1/reports/stock-overview/export?format=pdf&date_from={REPORT_DATE_FROM}"
            f"&date_to={REPORT_DATE_TO}&category={REPORT_GENERAL_CATEGORY}",
            "wms_stock_overview_2026-03-14.pdf",
        ),
        ("/api/v1/reports/surplus/export?format=pdf", "wms_surplus_2026-03-14.pdf"),
        (
            "/api/v1/reports/transactions/export?format=pdf"
            "&article_id={article_yellow_id}&date_from=2026-03-01&date_to=2026-03-31",
            "wms_transactions_2026-03-14.pdf",
        ),
    ],
)
def test_pdf_export_endpoints_return_downloads_with_expected_content_type(
    client,
    reports_data,
    path,
    expected_filename,
):
    response = client.get(
        path.format(article_yellow_id=reports_data["article_yellow_id"]),
        headers=_admin_headers(client, reports_data),
    )

    assert response.status_code == 200
    assert response.mimetype == "application/pdf"
    assert "attachment;" in response.headers["Content-Disposition"]
    assert _extract_filename(response.headers["Content-Disposition"]) == expected_filename
    assert response.data.startswith(b"%PDF")


def test_stock_overview_pdf_export_uses_landscape_title_subtitles_and_generic_rows(
    app,
    reports_data,
    monkeypatch,
):
    captured: dict[str, object] = {}

    def fake_build_pdf(**kwargs):
        captured.update(kwargs)
        return b"%PDF-stock"

    monkeypatch.setattr(report_service, "_build_pdf", fake_build_pdf)

    with app.app_context():
        content, filename, mimetype = report_service.export_stock_overview(
            export_format="pdf",
            date_from=REPORT_DATE_FROM,
            date_to=REPORT_DATE_TO,
            category=REPORT_GENERAL_CATEGORY,
        )

    assert content == b"%PDF-stock"
    assert filename == "wms_stock_overview_2026-03-14.pdf"
    assert mimetype == "application/pdf"
    assert captured["title"] == "Stock Overview"
    assert captured["landscape_mode"] is True
    assert captured["subtitle_lines"] == [
        "Date range: 2026-03-01 to 2026-03-31",
        "Exported at: 2026-03-14T12:00:00+00:00",
    ]
    assert "STOQIO" not in captured["title"]
    assert all("STOQIO" not in line for line in captured["subtitle_lines"])
    rows_by_article = {row[0]: row for row in captured["rows"]}
    assert rows_by_article["REP13-001"][11] == "YELLOW"
    assert rows_by_article["REP13-002"][9] == "∞"


def test_surplus_pdf_export_uses_portrait_title_timestamp_and_expected_rows(
    app,
    reports_data,
    monkeypatch,
):
    captured: dict[str, object] = {}

    def fake_build_pdf(**kwargs):
        captured.update(kwargs)
        return b"%PDF-surplus"

    monkeypatch.setattr(report_service, "_build_pdf", fake_build_pdf)

    with app.app_context():
        content, filename, mimetype = report_service.export_surplus_report(
            export_format="pdf",
        )

    assert content == b"%PDF-surplus"
    assert filename == "wms_surplus_2026-03-14.pdf"
    assert mimetype == "application/pdf"
    assert captured["title"] == "Surplus List"
    assert captured["landscape_mode"] is False
    assert captured["subtitle_lines"] == ["Exported at: 2026-03-14T12:00:00+00:00"]
    assert all("STOQIO" not in line for line in captured["subtitle_lines"])
    rows_by_article = {row[0]: row for row in captured["rows"]}
    assert rows_by_article["REP13-001"][4] == "3.0 rep13_kg"


def test_transaction_pdf_export_uses_landscape_title_date_range_timestamp_and_expected_rows(
    app,
    reports_data,
    monkeypatch,
):
    captured: dict[str, object] = {}

    def fake_build_pdf(**kwargs):
        captured.update(kwargs)
        return b"%PDF-transactions"

    monkeypatch.setattr(report_service, "_build_pdf", fake_build_pdf)

    with app.app_context():
        content, filename, mimetype = report_service.export_transaction_log(
            export_format="pdf",
            article_id=reports_data["article_yellow_id"],
            date_from=REPORT_DATE_FROM,
            date_to=REPORT_DATE_TO,
        )

    assert content == b"%PDF-transactions"
    assert filename == "wms_transactions_2026-03-14.pdf"
    assert mimetype == "application/pdf"
    assert captured["title"] == "Transaction Log"
    assert captured["landscape_mode"] is True
    assert captured["subtitle_lines"] == [
        "Date range: 2026-03-01 to 2026-03-31",
        "Exported at: 2026-03-14T12:00:00+00:00",
    ]
    assert all("STOQIO" not in line for line in captured["subtitle_lines"])
    assert [row[3] for row in captured["rows"]] == [
        "OUTBOUND",
        "SURPLUS_CONSUMED",
        "STOCK_CONSUMED",
        "STOCK_RECEIPT",
    ]
    assert captured["rows"][0][4] == "-1.0 rep13_kg"


@pytest.mark.parametrize(
    "path",
    [
        f"/api/v1/reports/stock-overview/export?format=xlsx&date_from={REPORT_DATE_FROM}&date_to={REPORT_DATE_TO}",
        f"/api/v1/reports/stock-overview/export?format=pdf&date_from={REPORT_DATE_FROM}&date_to={REPORT_DATE_TO}",
        "/api/v1/reports/surplus/export?format=xlsx",
        "/api/v1/reports/surplus/export?format=pdf",
        f"/api/v1/reports/transactions/export?format=xlsx&article_id=1&date_from={REPORT_DATE_FROM}&date_to={REPORT_DATE_TO}",
        f"/api/v1/reports/transactions/export?format=pdf&article_id=1&date_from={REPORT_DATE_FROM}&date_to={REPORT_DATE_TO}",
    ],
)
def test_manager_receives_403_on_all_export_endpoints(client, reports_data, path):
    response = client.get(path, headers=_manager_headers(client, reports_data))

    assert response.status_code == 403
    assert response.get_json()["error"] == "FORBIDDEN"


# ---------------------------------------------------------------------------
# RBAC read-access tests
# ---------------------------------------------------------------------------


@pytest.mark.parametrize("role_name", ["admin", "manager"])
def test_admin_and_manager_can_access_non_export_reports_endpoints(
    client,
    reports_data,
    role_name,
):
    headers = (
        _admin_headers(client, reports_data)
        if role_name == "admin"
        else _manager_headers(client, reports_data)
    )
    paths = [
        "/api/v1/reports/stock-overview"
        f"?date_from={REPORT_DATE_FROM}&date_to={REPORT_DATE_TO}&category={REPORT_GENERAL_CATEGORY}",
        "/api/v1/reports/surplus",
        f"/api/v1/reports/transactions?article_id={reports_data['article_yellow_id']}&page=1&per_page=5",
        "/api/v1/reports/statistics/top-consumption?period=month",
        "/api/v1/reports/statistics/movement?range=6m",
        "/api/v1/reports/statistics/reorder-summary",
        "/api/v1/reports/statistics/personal-issuances",
    ]

    for path in paths:
        response = client.get(path, headers=headers)
        assert response.status_code == 200, (path, response.get_json())


# ---------------------------------------------------------------------------
# Statistics tests
# ---------------------------------------------------------------------------


def test_top_consumption_statistics_returns_seeded_leader(client, reports_data):
    response = client.get(
        "/api/v1/reports/statistics/top-consumption?period=month",
        headers=_manager_headers(client, reports_data),
    )

    assert response.status_code == 200, response.get_json()
    payload = response.get_json()
    assert payload["period"] == "month"
    assert payload["date_from"] == "2026-03-01"
    assert payload["date_to"] == "2026-03-14"
    assert payload["items"][0]["article_no"] == "REP13-099"
    assert payload["items"][0]["outbound"] == 1000.0


def test_movement_statistics_return_seeded_month_delta(client, reports_data):
    response = client.get(
        "/api/v1/reports/statistics/movement?range=6m",
        headers=_manager_headers(client, reports_data),
    )

    assert response.status_code == 200, response.get_json()
    payload = response.get_json()
    assert payload["range"] == "6m"
    assert payload["granularity"] == "month"
    assert len(payload["items"]) == 6
    assert payload["note"] == (
        "Quantities are summed across all units of measure. "
        "This chart shows trends, not precise totals."
    )

    march_bucket = next(item for item in payload["items"] if item["bucket"] == "2026-03")
    assert march_bucket["inbound"] == pytest.approx(
        reports_data["baseline_march"]["inbound"] + float(SEEDED_MARCH_INBOUND)
    )
    assert march_bucket["outbound"] == pytest.approx(
        reports_data["baseline_march"]["outbound"] + float(SEEDED_MARCH_OUTBOUND)
    )


def test_reorder_summary_statistics_return_locked_uppercase_status_counts(client, reports_data):
    response = client.get(
        "/api/v1/reports/statistics/reorder-summary",
        headers=_manager_headers(client, reports_data),
    )

    assert response.status_code == 200, response.get_json()
    payload = response.get_json()
    counts = {item["reorder_status"]: item["count"] for item in payload["items"]}
    assert set(counts) == {"RED", "YELLOW", "NORMAL"}
    assert counts["RED"] == reports_data["baseline_reorder"]["RED"] + SEEDED_REORDER_INCREMENT["RED"]
    assert counts["YELLOW"] == reports_data["baseline_reorder"]["YELLOW"] + SEEDED_REORDER_INCREMENT["YELLOW"]
    assert counts["NORMAL"] == reports_data["baseline_reorder"]["NORMAL"] + SEEDED_REORDER_INCREMENT["NORMAL"]


def test_personal_issuances_statistics_return_current_year_seeded_row(client, reports_data):
    response = client.get(
        "/api/v1/reports/statistics/personal-issuances",
        headers=_manager_headers(client, reports_data),
    )

    assert response.status_code == 200, response.get_json()
    payload = response.get_json()
    assert payload["year"] == 2026
    seeded_row = next(
        item
        for item in payload["items"]
        if item["employee_name"] == "Ivo Ivic" and item["article_no"] == "REP13-004"
    )
    assert seeded_row["job_title"] == "Reports Painter"
    assert seeded_row["quantity_issued"] == 4.0
    assert seeded_row["quota"] == 10.0
    assert seeded_row["remaining"] == 6.0
    assert seeded_row["uom"] == "rep13_kom"
    assert seeded_row["quota_uom"] == "rep13_kom"
